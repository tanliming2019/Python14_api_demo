# _*_ coding:utf-8 _*_
# @time     :  3:36
# @Author   : lemon_Liming
# @Email    : 1104197970@qq.com
# @File     : do_excle.
#1：写一个类 类的作用是完成Excel数据的读写 新建表单的操作
#函数一：读取指定表单的数据， #有一个列表row_list,把每一行的每一个单元格的数据存到row_list1里面去。
#每一行都有 一个单独的row_list2 [[1,2,3],[4,5,6]] #每一行数据读取完毕后，把row_list存到大列表all_row_list
#函数二：在指定的单元格写入指定的数据，并保存到当前Excel
#函数三：新建一个Excel

from openpyxl import workbook #新建数据
from openpyxl import load_workbook#读取数据
from week_7.class_0227.homework.homework_1 import ReadLoging
logger=ReadLoging()


class DoExcel:

    '''类的作用是完成Excle的读写，新建表单的操作'''

    #嵌套字典
    logger.info('开始读取数据')
    def reserve_excle2(self, file_name, sheet_name):
        wb = load_workbook(file_name)  # 打开工作簿
        sheet = wb[sheet_name]#定位表单

        all_row_list = []  # 存所有数据的空列表
        for i in range(3,sheet.max_row+1):
            row_data ={} # 存每一行数据的空字典
            row_data['A'] = sheet.cell(i,1).value
            row_data['B'] = sheet.cell(i,2).value
            row_data['C'] = sheet.cell(i,3).value
            row_data['D'] = sheet.cell(i,4).value
            row_data['E'] = sheet.cell(i,5).value
            row_data['F'] = sheet.cell(i,6).value
            all_row_list.append(row_data)
        logger.info('数据读取完毕')
        return all_row_list  # 返回大列表的数据


    def write_excel(self,file_name,sheet_name,row,col,value):
        '''在指定的单元格写入指定的数据，并保存到当前的Excel'''
        wb=load_workbook(file_name)#打开工作簿
        sheet=wb[sheet_name]#定位表单

        logger.info('开始写入数据')
        try:
            sheet.cell(row,col).value=value
            wb.save(file_name)
            wb.close()#每次操作完要关闭进程
        except Exception as e:
            logger.error(e)

        logger.info('数据写入完毕')
    #
    def creat_excel(self,file_name,sheet_name):
        '''新建一个Excel'''
        wb=workbook.Workbook()#新建一个Excel文件
        wb.create_sheet(sheet_name)#新建一个表单
        wb.save(file_name)#保存文件名为

if __name__ == '__main__':
    # DoExcel().creat_excel('python17.xlsx','lemo')
    # DoExcel().write_excel('python17.xlsx','lemo',5,2,'信Python得永生')
    result=DoExcel().reserve_excle2('python22.xlsx','Sheet1')
    print(result)