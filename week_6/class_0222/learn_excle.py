# _*_ coding:utf-8 _*_
# @time     :  0:01
# @Author   : lemon_Liming
# @Email    : 1104197970@qq.com
# @File     : learn_excle.py
from openpyxl import load_workbook
#打开工作簿
wb=load_workbook('python14.xlsx')
#定位表单
sheet=wb['Sheet1']
#定位单元格
res=sheet.cell(4,1).value#在指定单元格写入sheet.cell(3,2,'作者：李白')内容
print(res)
# 打印第六行全部内容
for i in range(1,sheet.max_column+1):#+1 range取左不取右
    res=sheet.cell(column=i,row=6).value
    print(res)
#打印全部
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        res=sheet.cell(row=i,column=j).value
        if res !=None:
            print(res)
#写数据

wb.save('python15.xlsx')


# #新建工作簿
# from openpyxl import workbook
# wb=workbook.Workbook()
wb.save('python16.xlsx')