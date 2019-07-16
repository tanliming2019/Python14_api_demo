# _*_ coding:utf-8 _*_
# @time     :  3:42
# @Author   : lemon_Liming
# @Email    : 1104197970@qq.com
# @File     : homework_2.py
# 2：思考：分别将我们学过的数据类型  int float boolean str list tuple dict 写到每个单元格里面，
# 观察，你通过openpyxl操作后拿到的数据分别是是什么类型。

from openpyxl import  load_workbook
from openpyxl import workbook

from openpyxl import workbook #新建数据
from openpyxl import load_workbook#读取数据
wb=workbook.Workbook()#新建一个Excel文件
wb.create_sheet('Sheet1')#新建一个表单
wb.save('python18.xlsx')#保存文件名为
wb=load_workbook('python18.xlsx')#打开工作簿
sheet=wb['Sheet1']#定位表单
#写数据进Sheet1表单
res_2=sheet.cell(2,1).value='柠檬'
res_3=sheet.cell(2,2).value=101
res_4=sheet.cell(2,3).value=0.01
res_5=sheet.cell(2,4).value=True
res_6=sheet.cell(2,5).value='[1,2]'
res_7=sheet.cell(2,5).value='(1,2)'
res_8=sheet.cell(2,5).value='{1:2}'
wb.save('python18.xlsx')
wb.close()#每次操作完要关闭进程
print(type(res_2))
print(type(res_3))
print(type(res_4))
print(type(res_5))
print(type(res_6))
print(type(res_7))
print(type(res_8))