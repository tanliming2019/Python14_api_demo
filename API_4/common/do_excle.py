# _*_ coding:utf-8 _*_
# @time     :  22:43
# @Author   : lemon_Liming
# @Email    : 1104197970@qq.com
# @File     : do_excle.py

from openpyxl import load_workbook
from API_4.common import project_path
from API_4.common.read_config import ReadConfig
class DoExcle:
    '''该类完成测试数据的读取以及测试数据的写回'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name#excle工作簿文件名或地址
        self.sheet_name=sheet_name#表单名

    def read_data(self,section):
        '''从excle读取数据，有返回值'''
        #从配置文件里面控制读取那些用例
        case_id=ReadConfig(project_path.conf_path).get_data(section,'case_id')#section是配置里面的片段名，可以根据你的指定来测试具体的用例
        try:
            wb=load_workbook(self.file_name)#打开excle工作簿
            global sheet#声明sheet为全局变量
            sheet=wb[self.sheet_name]#定位表单
        except Exception as e:
            print('发生错误{}'.format(e))

        #每一行数据要在一起[]{}
        #每一行数据存到一个空间里
        #开始读数据
        tel=self.get_tel()
        test_data=[]#存放所有数据
        for i in range(2,sheet.max_row+1):
            row_data={}#存放每一行的数据
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Module']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Url'] = sheet.cell(i, 4).value
            row_data['Method']=sheet.cell(i,5).value
            if sheet.cell(i,6).value.find('tel')!=-1:#注意find这个方法的使用以及返回值
                row_data['Params']=sheet.cell(i,6).value.replace('tel',str(tel))
                self.update_tel(int(tel)+1)  #注意 这里的tel获取的是字符串类型
            else:
                row_data['Params']=sheet.cell(i,6).value
            row_data['ExpectedResult']=sheet.cell(i,7).value
            test_data.append(row_data)#把每一行的数据放到一个列表
        wb.close()

        #完成用例的可配置化
        final_data=[]   #空列表  储存最终的测试用例数据
        if case_id=='all':  #如果case_id==all,那就获取所有测试用例
            final_data=test_data   #把测试用例赋值给final_data这个变量
        else:  #否则，如果是列表，那就获取列表里面指定id的用例数据
            for i in case_id:
                final_data.append(test_data[i-1])
        return final_data

    def get_tel(self):
        '''获取存在excle里面的电话号码'''
        wb=load_workbook(self.file_name) #打开工作簿
        sheet=wb['tel'] #定位表单
        wb.close()
        return sheet.cell(1,2).value

    def update_tel(self,new_tel):
        '''写回手机号码'''
        wb=load_workbook(self.file_name) #打开工作簿
        sheet=wb['tel']  # 定位表单
        sheet.cell(1,2,new_tel)
        wb.save(self.file_name)
        wb.close()

    def write_back(self,row,col,value):
        '''写回测试结果到excle中'''
        wb=load_workbook(self.file_name)#打开工作簿
        sheet=wb[self.sheet_name]#定位表单
        sheet.cell(row,col).value=value#写入数据
        wb.save(self.file_name)
        wb.close()



if __name__ == '__main__':
    file_name=project_path.case_path
    sheet_name='add_loan'
    test_data=DoExcle(file_name,sheet_name).read_data('AddLoanCASE')
    print(test_data)
    # res=DoExcle(project_path.case_path,'tel').get_tel()
    # print(res)













