# _*_ coding:utf-8 _*_
# @time     :  22:43
# @Author   : lemon_Liming
# @Email    : 1104197970@qq.com
# @File     : do_excle.py

from openpyxl import load_workbook
from API_1.common import project_path
from API_2_作业练习.common.get_caseid import DataType
class DoExcle:
    '''该类完成测试数据的读取以及测试数据的写回'''
    def __init__(self,file_name,sheet_name,result):
        self.file_name=file_name#excle工作簿文件名或地址
        self.sheet_name=sheet_name#表单名
        self.result = result

    def read_data(self):
        '''从excle读取数据，有返回值'''
        try:
            global wb
            wb=load_workbook(self.file_name)#打开excle工作簿
            global sheet#声明sheet为全局变量
            sheet=wb[self.sheet_name]#定位表单
        except Exception as e:
            print('发生错误{}'.format(e))

        #每一行数据要在一起[]{}
        #每一行数据存到一个空间里
        #开始读数据
        test_data=[]#存放所有数据



        for i in self.result:  #[3,6,9]  result必须大于等于2
            i=i+1
            row_data={}#存放每一行的数据
            row_data['Caseid']=sheet.cell(i,1).value
            row_data['Module']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Url'] = sheet.cell(i, 4).value
            row_data['Method']=sheet.cell(i,5).value
            row_data['Params']=sheet.cell(i,6).value
            row_data['ExpectedResult']=sheet.cell(i,7).value
            test_data.append(row_data)#把每一行的数据放到一个列表
        wb.close()
        return test_data

    def write_back(self,row,col,value):
        '''写回测试结果到excle中'''
        wb=load_workbook(self.file_name)#打开工作簿
        sheet=wb[self.sheet_name]#定位表单
        sheet.cell(row,col).value=value#写入数据
        wb.save(self.file_name)
        wb.close()



if __name__ == '__main__':
    file_name=project_path.case_path
    sheet_name='test_case'
    result=DataType('case_id.conf').get_data('Case','id')
    test_data=DoExcle(file_name,sheet_name,result).read_data()
    print(test_data)













